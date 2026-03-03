#!/usr/bin/env python3
"""
Amazon 商品数据爬虫
爬取 Amazon 商品数据，生成 Excel + Markdown 报告

作者：Elden (@zuokun300)
许可：MIT
"""

import json
import re
import urllib.request
import sys
import os
from datetime import datetime

# 尝试导入依赖
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("❌ 缺少依赖：openpyxl")
    print("请运行：pip3 install openpyxl requests --break-system-packages")
    sys.exit(1)

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("❌ 缺少依赖：requests")
    print("请运行：pip3 install openpyxl requests --break-system-packages")
    sys.exit(1)


# ============= 配置区域 =============

# Apify API Key（从环境变量读取）
# 使用前请设置：export APIFY_API_KEY="apify_api_xxxxx"
APIFY_TOKEN = os.getenv("APIFY_API_KEY")

if not APIFY_TOKEN:
    print("❌ 错误：未设置 APIFY_API_KEY 环境变量")
    print("请运行：export APIFY_API_KEY=\"apify_api_xxxxx\"")
    print("获取 API Key: https://console.apify.com/account/integrations")
    sys.exit(1)

# 搜索关键词
DEFAULT_KEYWORDS = ["women fashion shoes"]

# 最大商品数量
MAX_PRODUCTS = 50

# 输出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# ============= 核心函数 =============

def extract_asin(url):
    """从 Amazon URL 提取 ASIN"""
    match = re.search(r'/dp/([A-Z0-9]{10})', url)
    return match.group(1) if match else None


def fetch_apify_data(keywords, max_products=50):
    """
    从 Apify 获取 Amazon 数据
    
    Args:
        keywords: 搜索关键词列表
        max_products: 最大商品数量
    
    Returns:
        商品数据列表
    """
    print(f"🔍 开始爬取 Amazon 数据...")
    print(f"   关键词：{', '.join(keywords)}")
    print(f"   目标数量：{max_products} 个商品")
    
    # 启动 Apify Actor
    start_url = f"https://www.amazon.com/s?k={keywords[0].replace(' ', '+')}"
    
    apify_url = "https://api.apify.com/v2/acts/apify~web-scraper/runs"
    params = {
        "token": APIFY_TOKEN
    }
    
    payload = {
        "startUrls": [{"url": start_url}],
        "linkSelector": "a.a-link-normal",
        "pageFunction": "function pageFunction() { return { title: document.title, url: window.location.href }; }"
    }
    
    try:
        # 启动爬取任务
        response = requests.post(apify_url, params=params, json=payload)
        run_data = response.json()
        
        if "error" in run_data:
            print(f"❌ 启动 Apify 失败：{run_data['error']['message']}")
            return []
        
        run_id = run_data["data"]["id"]
        dataset_id = run_data["data"]["defaultDatasetId"]
        
        print(f"✅ 爬取任务已启动：{run_id}")
        print(f"⏳ 等待爬取完成...")
        
        # 等待爬取完成（最多等待 5 分钟）
        import time
        for i in range(30):
            time.sleep(10)
            
            # 检查状态
            status_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
            status_response = requests.get(status_url, params=params)
            status_data = status_response.json()
            
            status = status_data["data"]["status"]
            if status == "SUCCEEDED":
                print(f"✅ 爬取完成！")
                break
            elif status == "FAILED":
                print(f"❌ 爬取失败")
                return []
            
            print(f"   当前状态：{status} ({i*10}s)")
        
        # 获取数据
        dataset_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
        data_response = requests.get(dataset_url, params=params)
        data = data_response.json()
        
        print(f"📦 获取到 {len(data)} 条原始数据")
        return data
        
    except Exception as e:
        print(f"❌ 爬取失败：{e}")
        return []


def generate_report(data, keywords):
    """
    生成商品报告
    
    Args:
        data: 原始数据
        keywords: 搜索关键词
    
    Returns:
        整理后的商品列表
    """
    products = []
    
    for item in data:
        title = item.get('title', '')
        debug = item.get('#debug', {})
        loaded_url = debug.get('loadedUrl', '')
        
        # 跳过搜索页面
        if '/s?' in loaded_url:
            continue
        
        asin = extract_asin(loaded_url)
        if not asin:
            continue
        
        # 从标题提取品牌
        brand = "Unknown"
        title_lower = title.lower()
        if "adidas" in title_lower:
            brand = "Adidas"
        elif "nike" in title_lower:
            brand = "Nike"
        elif "adokoo" in title_lower:
            brand = "Adokoo"
        elif "odoly" in title_lower:
            brand = "ODOLY"
        elif "lucky step" in title_lower:
            brand = "LUCKY STEP"
        elif "new balance" in title_lower:
            brand = "New Balance"
        
        # 判断鞋子类型
        shoe_type = "Sneaker"
        if "boot" in title_lower:
            shoe_type = "Boot"
        elif "sandal" in title_lower:
            shoe_type = "Sandal"
        elif "heel" in title_lower:
            shoe_type = "Heel"
        elif "clog" in title_lower or "mule" in title_lower:
            shoe_type = "Clog"
        
        # 清理标题
        clean_title = title.replace('Amazon.com | ', '').replace(' | Shoes', '').replace(' | Fashion Sneakers', '')
        
        products.append({
            'title': clean_title,
            'asin': asin,
            'brand': brand,
            'type': shoe_type,
            'url': loaded_url,
            'image': f"https://m.media-amazon.com/images/I/61uA2UVnYWL._AC_UL320_.jpg"
        })
    
    # 去重
    seen = set()
    unique_products = []
    for p in products:
        if p['asin'] not in seen:
            seen.add(p['asin'])
            unique_products.append(p)
    
    return unique_products[:max_products]


def generate_excel(products, filepath, keywords):
    """生成 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品数据"
    
    # 表头
    headers = ["排名", "商品名称", "品牌", "类型", "ASIN", "价格", "图片链接", "Amazon 链接"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 数据行
    for i, p in enumerate(products, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=p['title'])
        ws.cell(row=i, column=3, value=p['brand'])
        ws.cell(row=i, column=4, value=p['type'])
        ws.cell(row=i, column=5, value=p['asin'])
        ws.cell(row=i, column=6, value="待爬取")
        
        img_cell = ws.cell(row=i, column=7, value=p['image'])
        img_cell.hyperlink = p['image']
        img_cell.style = "Hyperlink"
        
        url_cell = ws.cell(row=i, column=8, value=p['url'])
        url_cell.hyperlink = p['url']
        url_cell.style = "Hyperlink"
    
    # 调整列宽
    column_widths = [8, 60, 15, 12, 15, 12, 50, 80]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(products)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # 统计 Sheet
    ws_stats = wb.create_sheet(title="统计摘要")
    
    brand_count = {}
    for p in products:
        brand = p['brand']
        brand_count[brand] = brand_count.get(brand, 0) + 1
    
    ws_stats.cell(row=1, column=1, value="品牌分布统计")
    ws_stats.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws_stats.cell(row=3, column=1, value="品牌")
    ws_stats.cell(row=3, column=2, value="商品数量")
    ws_stats.cell(row=3, column=3, value="占比")
    
    total = len(products)
    for i, (brand, count) in enumerate(sorted(brand_count.items(), key=lambda x: x[1], reverse=True), 4):
        ws_stats.cell(row=i, column=1, value=brand)
        ws_stats.cell(row=i, column=2, value=count)
        ws_stats.cell(row=i, column=3, value=f"{count/total*100:.1f}%")
    
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 10
    
    wb.save(filepath)
    return filepath


def generate_markdown(products, filepath, keywords):
    """生成 Markdown 报告"""
    md = f"# Amazon 商品数据报告\n\n"
    md += f"## 数据概览\n\n"
    md += f"- **抓取时间**: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    md += f"- **关键词**: {', '.join(keywords)}\n"
    md += f"- **商品数量**: {len(products)}\n\n"
    
    md += f"## 爆款商品列表\n\n"
    md += "| 排名 | 商品 | 品牌 | 类型 | ASIN | 图片 |\n"
    md += "|------|------|------|------|------|------|\n"
    
    for i, p in enumerate(products, 1):
        title_short = p['title'][:40] + "..." if len(p['title']) > 40 else p['title']
        md += f"| {i} | {title_short} | {p['brand']} | {p['type']} | {p['asin']} | ![img]({p['image']}) |\n"
    
    md += f"\n## 商品详情链接\n\n"
    for i, p in enumerate(products, 1):
        md += f"{i}. **{p['title']}** - [Amazon 链接]({p['url']})\n"
    
    # 品牌统计
    brand_count = {}
    for p in products:
        brand = p['brand']
        brand_count[brand] = brand_count.get(brand, 0) + 1
    
    md += f"\n## 品牌分布\n\n"
    md += "| 品牌 | 商品数量 | 占比 |\n"
    md += "|------|---------|------|\n"
    
    total = len(products)
    for brand, count in sorted(brand_count.items(), key=lambda x: x[1], reverse=True):
        md += f"| {brand} | {count} | {count/total*100:.1f}% |\n"
    
    with open(filepath, "w", encoding='utf-8') as f:
        f.write(md)
    
    return filepath


# ============= 主函数 =============

def main(keywords=None, max_products=50, output_dir=None):
    """主函数"""
    if keywords is None:
        keywords = DEFAULT_KEYWORDS
    
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    # 获取数据
    data = fetch_apify_data(keywords, max_products)
    
    if not data:
        print("❌ 未获取到数据")
        return
    
    # 生成报告
    products = generate_report(data, keywords)
    print(f"✅ 整理出 {len(products)} 个商品")
    
    # 生成文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"amazon-data-{timestamp}.xlsx")
    md_path = os.path.join(output_dir, f"amazon-report-{timestamp}.md")
    
    generate_excel(products, excel_path, keywords)
    generate_markdown(products, md_path, keywords)
    
    print(f"\n✅ 报告已生成：")
    print(f"   📊 Excel: {excel_path}")
    print(f"   📝 Markdown: {md_path}")
    
    # 显示前 5 个商品
    print(f"\n🏆 Top 5 爆款商品:")
    for i, p in enumerate(products[:5], 1):
        print(f"   {i}. {p['brand']} - {p['title'][:50]}...")
    
    return products


# ============= 命令行入口 =============

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Amazon 商品数据爬虫")
    parser.add_argument("-k", "--keyword", nargs="+", default=DEFAULT_KEYWORDS, help="搜索关键词")
    parser.add_argument("-n", "--number", type=int, default=MAX_PRODUCTS, help="最大商品数量")
    parser.add_argument("-o", "--output", default=OUTPUT_DIR, help="输出目录")
    
    args = parser.parse_args()
    
    main(keywords=args.keyword, max_products=args.number, output_dir=args.output)
